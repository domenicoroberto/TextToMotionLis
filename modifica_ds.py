import os

import openpyxl
import pandas as pd
from openpyxl import load_workbook

def modifica_intestazione(file_path):
    # Carica il file Excel
    workbook = openpyxl.load_workbook(file_path)
    # Seleziona il primo foglio di lavoro
    sheet = workbook.active

    # Modifica la seconda cella della prima riga
    sheet.cell(row=1, column=2).value = 'hands'

    # Salva il file Excel con le modifiche
    workbook.save(file_path)

def process_xlsx_files(input_folder, output_folder):

    for root, dirs, files in os.walk(input_folder):
        for dir_name in dirs:
            input_dir_path = os.path.join(root, dir_name)
            output_dir_path = os.path.join(output_folder, dir_name)
            os.makedirs(output_dir_path, exist_ok=True)

            for file_name in os.listdir(input_dir_path):
                if file_name.endswith('.xlsx'):
                    xlsx_path = os.path.join(input_dir_path, file_name)
                    print(xlsx_path)
                    modifica_intestazione(xlsx_path)

                    #fill_empty_cells(xlsx_path)
                    '''
                    output_file_path = os.path.join(output_dir_path, file_name)
                    os.makedirs(output_dir_path, exist_ok=True)
                    wb.save(output_file_path)
'''
# Utilizzo dell'esempio:
input_folder = r"D:\Lis\dataset\keypoint_video_tagliati"
output_folder = r"D:\Lis\dataset\keypoint_video_tagliati2"

process_xlsx_files(input_folder, output_folder)